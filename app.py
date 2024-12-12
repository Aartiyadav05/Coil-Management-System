from flask import Flask, render_template, request, redirect, url_for, session, flash
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os

app = Flask(__name__)
app.secret_key = "your_secret_key"

# File Paths
USER_DATA_FILE = "user_data.xlsx"
COIL_DETAILS_FILE = "coil_details.xlsx"

# Initialize Excel Files if Not Exists
def init_excel_files():
    try:
        # User Data
        wb = openpyxl.load_workbook(USER_DATA_FILE)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Username", "Password"])
        wb.save(USER_DATA_FILE)
    try:
        # Coil Details
        wb = openpyxl.load_workbook(COIL_DETAILS_FILE)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([ "TDC","Coil Number","Thickness","Width","Quality Code","RM Batch","username","Timestamp"])
        wb.save(COIL_DETAILS_FILE)

init_excel_files()

@app.route('/')
def home():
    return redirect(url_for('login'))

# Login Page
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        wb = openpyxl.load_workbook(USER_DATA_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username and row[1] == password:
                session['username'] = username
                return redirect(url_for('options'))
        flash("Invalid username or password", "error")
    return render_template('login.html')

# Register Page
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        confirm_password = request.form['confirm_password']

        # Validations 
        if len(password) < 8 or not any(char.isupper() for char in password):
            flash("Password must be at least 8 characters long and contain at least one uppercase letter", "error")
            return redirect(url_for('register'))
        if password != confirm_password:
            flash("Passwords do not match", "error")
            return redirect(url_for('register'))

        wb = openpyxl.load_workbook(USER_DATA_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username:
                flash("Username already exists", "error")
                return redirect(url_for('register'))

        ws.append([username, password])
        wb.save(USER_DATA_FILE)
        flash("Registration successful. Please login.", "success")
        return redirect(url_for('login'))

    return render_template('register.html')

# Options Page
@app.route('/options')
def options():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('options.html')
 

          
@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    workbook=load_workbook(COIL_DETAILS_FILE)
    sheet=workbook.active
    if request.method == 'POST':
                               
        # Get data from the form
        tdc = request.form.get('tdc').upper()
        coil_number = request.form.get('coil_number').upper()
        thickness = request.form.get('thickness').upper()
        width = request.form.get('width','').strip().upper()
        quality_code = request.form.get('quality_code').upper()
        rm_batch = request.form.get('rm_batch').upper()
        username = session.get('username')
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S').upper()
         
        if username =="your_valid_username":
            session['username']=username
        
        if not (len(tdc) == 4 and tdc.isalnum()):
            flash("TDC must be 4 alphanumeric characters.")
            return redirect(url_for('dashboard'))

        # Coil Number Validation: Max 10 characters
        if len(coil_number) > 10:
            flash("Coil Number cannot exceed 10 characters.")
            return redirect(url_for('dashboard'))

        # Thickness Validation: Float
        try:
            float(thickness)
        except ValueError:
            flash("Thickness must be a valid float.")
            return redirect(url_for('dashboard'))
        
        try:
            width = float(width)
        except ValueError:
            flash('Width must be a float value.', 'error')
            return redirect(url_for('dashboard'))
    
        # Quality Code Validation: Max 6 characters
        if len(quality_code) > 6:
            flash("Quality Code cannot exceed 6 characters.")
            return redirect(url_for('dashboard'))

        # RM Batch Validation: Max 10 characters
        if len(rm_batch) > 10:
            flash("RM Batch cannot exceed 10 characters.")
            return redirect(url_for('dashboard'))
        
           # Check for duplicate TDC and Coil Number
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from the second row (skip headers)
            existing_tdc = row[0]
            existing_coil_number = row[1]
            if tdc == existing_tdc and coil_number == existing_coil_number:
                flash("Duplicate TDC and Coil Number entry detected.", "error")
                return redirect(url_for('dashboard'))
        
        try:
           
            # Check if the Excel file exists, if not create it
            if not os.path.exists(COIL_DETAILS_FILE):
                sheet = workbook.active
                sheet.title = "Coil Data"
                # Create headers
                sheet.append(["TDC", "Coil Number", "Thickness", "Quality Code", "RM Batch", "Width","username", "Timestamp"])
                workbook.save(COIL_DETAILS_FILE)
            
             
            # Load the workbook and append the data
            username = session.get('username')
            workbook = load_workbook(COIL_DETAILS_FILE)
            sheet = workbook.active
            sheet.append([tdc, coil_number, thickness,width, quality_code, rm_batch, username,timestamp])
            workbook.save(COIL_DETAILS_FILE)

            flash("Data saved successfully!", "success")
            return redirect(url_for('dashboard'))
        except Exception as e:
            flash(f"An error occurred: {e}", "error")
            return redirect(url_for('dashboard'))

    return render_template('dashboard.html')
       
 

@app.route('/tdc')
def tdc():
    if 'username' not in session:
        flash('Please log in to access this page.', 'error')
        return redirect(url_for('login'))

    unique_tdcs = set()
    try:
        # Load coil details Excel file
        workbook = load_workbook('coil_details.xlsx')
        sheet = workbook.active

        # Collect unique TDC values from the file
        for row in sheet.iter_rows(min_row=2, values_only=True):
            unique_tdcs.add(row[0])  # Assuming TDC is the first column

    except FileNotFoundError:
        flash('Coil details file not found.', 'error')

    return render_template('tdc.html', unique_tdcs=sorted(unique_tdcs))

 
 
@app.route('/show_report', methods=['GET'])
def show_report():
    if 'username' not in session:
        flash('Please log in to access this page.', 'error')
        return redirect(url_for('login'))

    selected_tdc = request.args.get('tdc')  # Retrieve TDC value from the form
    if not selected_tdc:
        flash('No TDC selected. Please try again.', 'error')
        return redirect(url_for('tdc'))

    data = []

    try:
        # Load coil details Excel file
        workbook = load_workbook('coil_details.xlsx')
        sheet = workbook.active

        # Collect matching rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == selected_tdc:  # Assuming TDC is in the first column
                data.append(row)

    except FileNotFoundError:
        flash('Coil details file not found.', 'error')
        return redirect(url_for('tdc'))

    # Fetch unique TDCs again for dropdown
    unique_tdcs = set(row[0] for row in sheet.iter_rows(min_row=2, values_only=True))

    return render_template(
        'tdc.html',
        unique_tdcs=unique_tdcs,
        data=data,
        selected_tdc=selected_tdc
    )


# Logout
@app.route('/logout')
def logout():
    session.pop('username', None)
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(debug=True)
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
#     # Dashboard Page
# @app.route('/dashboard', methods=['GET', 'POST'])
# def dashboard():
#     if 'username' not in session:
#         return redirect(url_for('login'))
#     if request.method == 'POST':
#         tdc = request.form['tdc']
#         coil_number = request.form['coil_number']
#         quality_code = request.form['quality_code']
#         width = request.form['width']
#         thickness = request.form['thickness']
#         rm_batch = request.form['rm_batch']

#         timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#         username = session['username']

#         wb = openpyxl.load_workbook(COIL_DETAILS_FILE)
#         ws = wb.active
#         ws.append([timestamp, username, coil_number, quality_code, tdc, width, thickness, rm_batch])
#         wb.save(COIL_DETAILS_FILE)

#         flash("Data saved successfully!", "success")
#     return render_template('dashboard.html')
    
    
#     @app.route('/dashboard', methods=['GET', 'POST'])
# def dashboard():
    

#     if request.method == 'POST':
#         try:
#             # Retrieve form data
#             tdc = request.form['tdc']
#             coil_number = request.form['coil_number']
#             thickness = request.form['thickness']
#             width = request.form['width']
#             quality_code = request.form['quality_code']
#             rm_batch = request.form['rm_batch']

#             # Validation checks
#             if len(tdc) != 4 or not tdc.isalnum():
#                 flash('TDC must be 4 characters and alphanumeric.', 'error')
#                 return redirect(url_for('dashboard'))

#             if len(coil_number) > 10 or not coil_number.isalpha():
#                 flash('Coil Number must contain up to 10 alphabetic characters.', 'error')
#                 return redirect(url_for('dashboard'))

#             try:
#                 thickness = float(thickness)
#             except ValueError:
#                 flash('Thickness must be a float value.', 'error')
#                 return redirect(url_for('dashboard'))

#             try:
#                 width = float(width)
#             except ValueError:
#                 flash('Width must be a float value.', 'error')
#                 return redirect(url_for('dashboard'))

#             if len(quality_code) != 6 or not quality_code.isalnum():
#                 flash('Quality Code must be exactly 6 alphanumeric characters.', 'error')
#                 return redirect(url_for('dashboard'))

#             if len(rm_batch) > 10:
#                 flash('RM Batch must not exceed 10 characters.', 'error')
#                 return redirect(url_for('dashboard'))

#             # Load or create coil_details.xlsx
#             wb, sheet = COIL_DETAILS_FILE()

#             # Check for duplicate TDC in the file
#             for row in sheet.iter_rows(min_row=2, values_only=True):
#                 if row[0] == tdc:
#                     flash('TDC already exists. Please enter a unique TDC.', 'error')
#                     return redirect(url_for('dashboard'))

#             # Save data to the Excel sheet
#             sheet.append([
#                 tdc,
#                 coil_number,
#                 thickness,
#                 width,
#                 quality_code,
#                 rm_batch,
#                 session.get('username', 'Anonymous'),  # Get username from session
#                 datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # Add a timestamp
#             ])
#             wb.save('coil_details.xlsx')  # Save the workbook

#             flash('Data saved successfully!', 'success')
#             return redirect(url_for('dashboard'))  # Redirect to the same page to clear the form
#         except Exception as e:
#             flash(f'An error occurred: {e}', 'error')
#             return redirect(url_for('dashboard'))

#     return render_template('dashboard.html')

    
# # Route for TDC Report Page
# @app.route('/tdc_report')
# def tdc_report():
#     # Load unique TDC numbers from the coil_details.xlsx file
#     workbook = openpyxl.load_workbook('coil_details.xlsx')
#     sheet = workbook.active

#     tdc_numbers = set()  # Use a set to store unique TDC numbers
#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         tdc_numbers.add(row[0])  # Assuming TDC numbers are in the first column

#     return render_template('tdc.html', tdc_numbers=sorted(tdc_numbers))


# # Route for Showing the Report based on the selected TDC number
# @app.route('/show_report', methods=['POST'])
# def show_report():
#     tdc_number = request.form['tdc']
#     report_data = []

#     # Load data from 'coil_details.xlsx'
#     workbook = openpyxl.load_workbook('coil_details.xlsx')
#     sheet = workbook.active

#     # Fetch data matching the TDC number
#     for<!DOCTYPE html> 
#         if row[0] == tdc_number:
#             report_data.append({
#                 'tdc': row[0],
#                 'coil_number': row[1],
#                 'thickness': row[2],
#                 'width': row[3],
#                 'quality_code': row[4],
#                 'rm_batch': row[5],
#                 'username':row[6],
#                 'timestamp': row[7],
#             })

#     return render_template('tdc.html', tdc_numbers=[], report_data=report_data)  # Pass data to template\
   
